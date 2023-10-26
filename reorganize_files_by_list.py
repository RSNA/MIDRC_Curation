#script for reorganizing directory with removal of folders for future use
#initial by jsho 

import os
import shutil
import sys
import openpyxl

def get_excel_values(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    values = []
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            values.append(cell.value)
    
    return values

def move_matching_folders(src, dst, tag_list):
    if not os.path.exists(dst):
        os.makedirs(dst)

    for foldername in os.listdir(src):
        folder_path = os.path.join(src, foldername)
        if os.path.isdir(folder_path) and foldername in tag_list:
            shutil.move(folder_path, os.path.join(dst, foldername))
            print(f"Moved {foldername} to {dst}")

def main():
    if len(sys.argv) != 4:
        print("Usage: python script.py <excel_file> <source_directory> <destination_directory>")
        sys.exit(1)

    excel_file = sys.argv[1]
    src_dir = sys.argv[2]
    dst_dir = sys.argv[3]

    if not os.path.isfile(excel_file):
        print(f"Error: {excel_file} does not exist or is not a file.")
        sys.exit(1)

    if not os.path.isdir(src_dir):
        print(f"Error: {src_dir} does not exist or is not a directory.")
        sys.exit(1)

    tag_list = get_excel_values(excel_file)
    move_matching_folders(src_dir, dst_dir, tag_list)

if __name__ == "__main__":
    main()
